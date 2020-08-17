from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import load_workbook
import support
from datetime import datetime
import pytest

wb = load_workbook('C:/iviyo/testcases.xlsx')
sheet = wb.active
max_row = sheet.max_row

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.implicitly_wait(1)

for i in range(2, max_row + 1):
    if str(sheet.cell(row=i, column=1).value).find('verify') != -1:
        test_row = i
        k = 0

    if str(sheet.cell(row=i, column=9).value) == 'Not Executed':

        if str(sheet.cell(row=i, column=7).value) == "get_url":
            url_outcome = support.get_url(driver,str(sheet.cell(row=i, column=8).value),str(sheet.cell(row=i, column=3).value))
            sheet['K' + str(i)] = url_outcome[0]
            sheet['L' + str(i)] = url_outcome[1]
            sheet['M' + str(i)] = url_outcome[2]
            sheet['I' + str(i)] = datetime.now()
            if url_outcome[1] == 'FAIL' and k == 0:
                sheet['N' + str(test_row)] = 'FAIL'
                k = 1
            else:
                if k != 1:
                    sheet['N' + str(test_row)] = 'PASS'

        if str(sheet.cell(row=i, column=7).value) == "search_element":
            search_outcome = support.search_element(driver, str(sheet.cell(row=i, column=8).value),str(sheet.cell(row=i, column=3).value), str(sheet.cell(row=i, column=10).value))
            sheet['K' + str(i)] = search_outcome[0]
            sheet['L' + str(i)] = search_outcome[1]
            sheet['M' + str(i)] = search_outcome[2]
            sheet['I' + str(i)] = datetime.now()
            if search_outcome[1] == 'FAIL':
                sheet['N' + str(test_row)] = 'FAIL'
                k = 1
            else:
                if k != 1:
                    sheet['N' + str(test_row)] = 'PASS'

        if str(sheet.cell(row=i, column=7).value) == "click_target_element":
            search_outcome = support.click_target_element(driver, str(sheet.cell(row=i, column=8).value),str(sheet.cell(row=i, column=3).value), str(sheet.cell(row=i, column=10).value))
            sheet['K' + str(i)] = search_outcome[0]
            sheet['L' + str(i)] = search_outcome[1]
            sheet['M' + str(i)] = search_outcome[2]
            sheet['I' + str(i)] = datetime.now()
            if search_outcome[1] == 'FAIL':
                sheet['N' + str(test_row)] = 'FAIL'
                k = 1
            else:
                if k != 1:
                    sheet['N' + str(test_row)] = 'PASS'

        if str(sheet.cell(row=i, column=7).value) == "single_input":
            search_outcome = support.single_input(driver, str(sheet.cell(row=i, column=8).value),str(sheet.cell(row=i, column=3).value), str(sheet.cell(row=i, column=10).value))
            sheet['K' + str(i)] = search_outcome[0]
            sheet['L' + str(i)] = search_outcome[1]
            sheet['M' + str(i)] = search_outcome[2]
            sheet['I' + str(i)] = datetime.now()
            if search_outcome[1] == 'FAIL':
                sheet['N' + str(test_row)] = 'FAIL'
                k = 1
            else:
                if k != 1:
                    sheet['N' + str(test_row)] = 'PASS'

        if str(sheet.cell(row=i, column=7).value) == "single_input_submit":
            search_outcome = support.single_input_submit(driver, str(sheet.cell(row=i, column=8).value),str(sheet.cell(row=i, column=3).value), str(sheet.cell(row=i, column=10).value))
            sheet['K' + str(i)] = search_outcome[0]
            sheet['L' + str(i)] = search_outcome[1]
            sheet['M' + str(i)] = search_outcome[2]
            sheet['I' + str(i)] = datetime.now()
            if search_outcome[1] == 'FAIL':
                sheet['N' + str(test_row)] = 'FAIL'
                k = 1
            else:
                if k != 1:
                    sheet['N' + str(test_row)] = 'PASS'

wb.save('C:/iviyo/testcases.xlsx')
support.close_process(driver)